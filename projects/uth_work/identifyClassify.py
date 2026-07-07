from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from classify import classify_response
import time
import subprocess


inp = input("Enter the identifier: ")
# inp = "4"

file_path = "sharedAnnotationsQA.xlsx"
# subprocess.run(f"mkdir info{inp}",shell=True, check=True) 
subprocess.run(f"mkdir -p info{inp}",shell=True, check=True) #-p causes creation only if it doesn't already exist
subprocess.run(f"cp {file_path} info{inp}/sharedAnnotationsQA{inp}.xlsx",shell=True, check=True)
# cpWb = load_workbook(f"info{inp}/sharedAnnotationsQA.xlsx", data_only=True)
# cpSheet = cpWb.active
# cpRows = list(cpSheet.iter_rows(
#     min_row=1,
#     # min_row=114,
#     max_col=6,
#     # max_row=3, #this processes only a few rows
#     values_only=True
# ))

file_path = f"info{inp}/sharedAnnotationsQA{inp}.xlsx"

workbook = load_workbook(file_path, data_only=True)
sheet = workbook.active
# Or use a specific sheet:
# sheet = workbook["Sheet1"]

# rows = list(sheet.iter_rows(values_only=True))
rows = list(sheet.iter_rows(
    min_row=1,
    # min_row=114,
    max_col=6,
    # max_row=3, #this processes only a few rows
    values_only=True
))

header = rows[0]




#extracting info on the sheet
# print(sheet.max_row)
# print(sheet.max_column)

#model mistake tracking
#  = [0,0,0,0] #order matches order valid, illogical, clarify, skip
mistakes = [[0 for col in range(5)] for row in range(5)]
matchingResponse = {
    f"{header[1]}": "valid",
    f"{header[2]}": "valid",
    f"{header[3]}": "illogical",
    f"{header[4]}": "clarify",
    f"{header[5]}": "skip"
}
matchingInd = {
    f"valid": 0,
    f"illogical": 1,
    f"clarify": 2,
    f"skip": 3,
    f"other": 4
}


# print(matchingInd.__contains__('othe'))

# print(matchingResponse["INCLUDE (falls under logical answer category)"])


#highlight info
# highlight_fills = [
#     "FF0000",  # red
#     "FFA500",  # orange
#     "FFFF00",  # yellow
#     "00FF00",  # lime green
#     "00B0F0",  # electric blue
# ]
highlight_fills = [
    "00FF00",  # lime green
    "FFA500",  # orange
    "FFFF00",  # yellow
    "00B0F0",  # electric blue
    "FF0000",  # red
]

fills = [
    PatternFill(start_color=color, end_color=color, fill_type="solid")
    for color in highlight_fills
]

#colorcoding key using header
for colInd in range(1, len(header)):  # starts at 1 to skip question column
    expected_category = matchingResponse[header[colInd]]
    fill_index = matchingInd[expected_category]

    sheet.cell(
        row=1,              # header row in Excel
        column=colInd + 1   # Excel columns are 1-based
    ).fill = fills[fill_index]

#for colorcoding font
from colorFont import terminal_fg

start = time.time()
checkpoint = start


for rowInd in range(1, len(rows)):  # row index in this list, not Excel row number
# for rowInd in range(1, 1):  # row index in this list, not Excel row number
# for rowInd in range(1, 3):  # row index in this list, not Excel row number
# for rowInd in range(115, len(rows)):  # row index in this list, not Excel row number
    #skipping rows used for prompt examples
    # if rowInd in range(50,60):
    #     continue

    row = rows[rowInd]
    
    print(f'\nTotal time elapsed: {(time.time()-start):.2f} seconds')
    # print(f"\nTime since last check: {(time.time()-checkpoint):.2f} seconds\n-------------------------------------------")
    checkpoint = time.time()

    for colInd in range(len(header)):
    # for colInd in range(6):
        mod = ""

        response = row[colInd] if colInd < len(row) else ""

        color = "FFFFFF"

        if colInd:  # not column 0
            if response == None:
                mod = ": blank"
            else:
                responseClassification = classify_response(row[0], response)

                # predictedColorInd = matchingInd[responseClassification]

                # mod = ': '+responseClassification
                
                if responseClassification == "nonresponsive": responseClassification = "illogical"
                if not matchingInd.__contains__(responseClassification): responseClassification = "other"

                #tracking mistakes of model
                #only mistakes
                # if responseClassification != matchingResponse[header[colInd]]:
                #     # print(colInd-1, matchingInd[matchingResponse[header[colInd]]])
                #     # mistakes[colInd-1][matchingInd[matchingResponse[header[colInd]]]] += 1
                #     print(colInd-1, matchingInd[responseClassification])
                #     mistakes[colInd-1][matchingInd[responseClassification]] += 1

                #correct inclusive
                # print(colInd-1, matchingInd[responseClassification])
                mod = ': '+terminal_fg(responseClassification, highlight_fills[matchingInd[responseClassification]])  #adding font color
                mistakes[colInd-1][matchingInd[responseClassification]] += 1 #modifying confusion matrix
                sheet.cell(row=rowInd+1, column=colInd+1).fill = fills[matchingInd[responseClassification]] #highlighting cell to identify correctness

                
                


        # print(f"{header[colInd]}: {response}{mod}")
        #safer print ig
        print(
            f"{header[colInd]}: {repr(str(response)[:300])}{mod}",
            flush=True
        )

        # break #skips everything while keeping all vars

workbook.save(file_path)
print(f'\n\n\nTotal time elapsed: {(time.time()-start):.2f} seconds')


#outputting mistakes
# mistakes[0][0] = -1
# mistakes[1][0] = -1
# mistakes[2][1] = -1
# mistakes[3][2] = -1
# mistakes[4][3] = -1

from toExcel5x5Headers import to5x5
to5x5(mistakes, ["yes", "no", "illogical", "clarify", "skip"], ["valid", "illogical", "clarify", "skip", "other"], output_file=f"info{inp}/mistakes_output{inp}.xlsx")

subprocess.run(f"cp -r info{inp} ~/port",shell=True, check=True)
