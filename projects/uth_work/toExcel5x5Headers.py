from openpyxl import Workbook
from openpyxl.styles import PatternFill


def to5x5(table, row_headers, col_headers, output_file="mistakes_output.xlsx"):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Mistakes"

    sheet.cell(row=1, column=1).value = "Actual \\ Predicted"

    #column headers
    for colInd, header in enumerate(col_headers, start=2):
        sheet.cell(row=1, column=colInd).value = header

    #row headers and values
    for rowInd, row_header in enumerate(row_headers, start=2):
        sheet.cell(row=rowInd, column=1).value = row_header

        for colInd, value in enumerate(table[rowInd - 2], start=2):
            sheet.cell(row=rowInd, column=colInd).value = value

    #highlight correct-match cells
    highlight_fill = PatternFill(
        start_color="00FF00",
        end_color="00FF00",
        fill_type="solid"
    )

    correct_cells = [
        (0, 0),  
        (1, 0),  
        (2, 1),  
        (3, 2),  
        (4, 3),  
    ]

    for rowInd, colInd in correct_cells:
        sheet.cell(row=rowInd + 2, column=colInd + 2).fill = highlight_fill

    wb.save(output_file)