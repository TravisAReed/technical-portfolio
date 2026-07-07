from openpyxl import load_workbook
from classify import classify_response
import time



"""
the original use of this was to copy an annotated excel document and create a new folder
then a confusion matrix was created and all the annotations were evaluated by the model
and had displayed what the model selected vs what they should have been for each one using
highlighting.

Also make sure header matches up with classifications in classify.py
"""

file_path = "sharedAnnotationsQA.xlsx" #location of original annotated file

workbook = load_workbook(file_path, data_only=True)
sheet = workbook.active
# #Or use a specific sheet:
# sheet = workbook["Sheet1"]

# rows = list(sheet.iter_rows(values_only=True))
rows = list(sheet.iter_rows(
    min_row=1,
    # min_row=114,
    max_col=6,
    # max_row=3, #this processes only a few rows
    values_only=True
))

header = rows[0]




#extracting info on the sheet
# print(sheet.max_row)
# print(sheet.max_column)

#model mistake tracking
#  = [0,0,0,0] #order matches order valid, illogical, clarify, skip
mistakes = [[0 for col in range(5)] for row in range(5)]
matchingResponse = {
    f"{header[1]}": "negative",
    f"{header[2]}": "negative",
    f"{header[3]}": "zero",
    f"{header[4]}": "one",
    f"{header[5]}": "more than one"
}
matchingInd = {
    f"negative": 0,
    f"zero": 1,
    f"one": 2,
    f"more than one": 3,
    f"other": 4
}


# print(matchingInd.__contains__('othe'))

# print(matchingResponse["INCLUDE (falls under logical answer category)"])

start = time.time()
checkpoint = start


for rowInd in range(1, len(rows)):  # row index in this list, not Excel row number
# for rowInd in range(115, len(rows)):  # row index in this list, not Excel row number
    #skipping rows used for prompt examples
    if rowInd in range(50,60):
        continue

    row = rows[rowInd]
    
    print(f'\nTotal time elapsed: {(time.time()-start):.2f} seconds')
    # print(f"\nTime since last check: {(time.time()-checkpoint):.2f} seconds\n-------------------------------------------")
    checkpoint = time.time()

    for colInd in range(len(header)):
    # for colInd in range(6):
        mod = ""

        response = row[colInd] if colInd < len(row) else ""

        if colInd:  # not column 0
            if response == None:
                mod = ": blank"
            else:
                responseClassification = classify_response(row[0], response)

                mod = ': '+responseClassification
                if not matchingInd.__contains__(responseClassification):
                    responseClassification = "other"

                #tracking mistakes of model
                #only mistakes
                # if responseClassification != matchingResponse[header[colInd]]:
                #     # print(colInd-1, matchingInd[matchingResponse[header[colInd]]])
                #     # mistakes[colInd-1][matchingInd[matchingResponse[header[colInd]]]] += 1
                #     print(colInd-1, matchingInd[responseClassification])
                #     mistakes[colInd-1][matchingInd[responseClassification]] += 1

                #correct inclusive
                # print(colInd-1, matchingInd[responseClassification])
                mistakes[colInd-1][matchingInd[responseClassification]] += 1
                


        #safer print with special characters
        print(
            f"{header[colInd]}: {repr(str(response)[:300])}{mod}",
            flush=True
        )

        # break #skips everything while keeping all vars

print(f'\n\n\nTotal time elapsed: {(time.time()-start):.2f} seconds')


#outputting mistakes
from toExcel5x5Headers import to5x5
#other is catch-all category if model makes a mistake (haven't seen that happen yet)
to5x5(mistakes, ["less than -1", "-1", "0", "1", "more than 1"], ["negative", "zero", "one", "more than one", "other"]) 